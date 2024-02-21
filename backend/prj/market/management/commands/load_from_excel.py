from django.core.management.base import BaseCommand
from prj.settings import DATA_DIR
from openpyxl import load_workbook
from market.models import Product, Category


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print('Start load data from excel...%s' % DATA_DIR)
        wb = load_workbook(DATA_DIR / 'price.xlsx')
        sheets = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheets[0])

        category = None
        for value in sheet.iter_rows(min_row=1, max_col=3, values_only=True):
            item = value[2]
            id = value[1]
            if id is None:
                category = Category()
                category.name = item
                category.save()
            else:
                if category:
                    product = Product()
                    product.name = item
                    product.category = category
                    product.save()
